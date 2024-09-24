import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

def remove_net_columns(df):
    """Elimina las columnas que comienzan con 'NET_'."""
    net_columns = [col for col in df.columns if col.startswith('NET_')]
    print(f"Columns to remove: {net_columns}")  # Depuración
    df = df.drop(columns=net_columns, errors='ignore')
    return df

def clean_sheet_name(sheet_name):
    """Limpia y ajusta el nombre de la hoja para cumplir con los requisitos de Excel."""
    sheet_name = sheet_name.replace('[', '.').replace(']', '').replace('{', '.').replace('}', '')
    sheet_name = re.sub(r'\.{2,}', '.', sheet_name)
    return sheet_name[:31]

# def format_headers(df):
#     """Renombra las columnas después de 'Verbatim' con COD_1, COD_2, etc."""
#     if 'Verbatim' in df.columns:
#         verbatim_index = df.columns.get_loc('Verbatim')
#         new_column_names = df.columns[:verbatim_index+1].tolist()
#         for i, col in enumerate(df.columns[verbatim_index+1:], start=1):
#             new_column_names.append(f'COD_{i}')
#         df.columns = new_column_names
#     return df

def format_headers(df):
    """Renombra las columnas a partir de la primera aparición de 'NET_' o 'Coding_' con COD_1, COD_2, etc."""
    # Palabras clave para buscar la primera columna objetivo
    keywords = ['NET_', 'Coding_']
    target_index = None

    # Buscar el índice de la primera columna que empiece con 'NET_' o 'Coding_'
    for keyword in keywords:
        keyword_indices = df.columns[df.columns.str.startswith(keyword)]
        if not keyword_indices.empty:
            first_index = df.columns.get_loc(keyword_indices[0])
            if target_index is None or first_index < target_index:
                target_index = first_index

    if target_index is not None:
        # Mantener los nombres de las columnas hasta la columna objetivo
        new_column_names = df.columns[:target_index].tolist()

        # Renombrar las columnas que siguen a la columna objetivo con 'COD_1', 'COD_2', etc.
        for i, col in enumerate(df.columns[target_index:], start=1):
            # Si la columna ya tiene un nombre 'COD_', mantenlo
            if 'COD_' in col:
                new_column_names.append(col)
            else:
                # Renombrar otras columnas con 'COD_' seguido del número
                new_column_names.append(f'COD_{i}')
        
        # Asignar los nuevos nombres de columna al DataFrame
        df.columns = new_column_names
    else:
        # Si no se encontró ninguna columna con los términos de búsqueda, no cambiar nada
        return df

    return df




def process_verbatim_code(sheet_data, writer):
    """Procesa la hoja 'Verbatim Code' y separa en hojas basadas en 'Question'."""
    if 'Question' in sheet_data.columns:
        questions = sheet_data['Question'].unique()
        for question in questions:
            df_filtered = sheet_data[sheet_data['Question'] == question]
            filtered_name = clean_sheet_name(str(question))
            if not df_filtered.empty:
                df_filtered.to_excel(writer, sheet_name=filtered_name, index=False)
    return writer

def separate_sheets(file_path, remove_net):
    print(f"Remove net columns: {remove_net}")
    output_file = os.path.join('uploads', '4_Codificacion_Easy_Code.xlsx')
    print(f"Output file path: {output_file}")

    sheets_failed = []
    os.makedirs('uploads', exist_ok=True)  # Asegura que la carpeta exista

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df = pd.read_excel(file_path, sheet_name=None)

            for sheet_name, data in df.items():
                if data.empty:
                    continue

                if remove_net:
                    print(f"Processing sheet: {sheet_name} - Removing NET_ columns")
                    data = remove_net_columns(data)

                if sheet_name == 'Verbatim Code':
                    writer = process_verbatim_code(data, writer)
                else:
                    clean_name = clean_sheet_name(sheet_name)
                    try:
                        data.to_excel(writer, sheet_name=clean_name, index=False)
                    except Exception as e:
                        sheets_failed.append(sheet_name)
                        print(f"Failed to write sheet {sheet_name}: {e}")

    except Exception as e:
        print(f"An error occurred: {e}")
        return None, ['File processing error']

    if sheets_failed:
        return None, sheets_failed

    if os.path.isfile(output_file):
        print(f"File successfully created: {output_file}")
    else:
        print(f"File creation failed: {output_file}")

    return output_file, None


def apply_formatting(output_file):
    """Aplica el formateo de encabezados a las hojas que contienen la columna 'Verbatim'."""
    temp_file = os.path.join('uploads', 'temp_formatting.xlsx')
    sheets_failed = []

    with pd.ExcelFile(output_file) as xls:
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'Verbatim' in df.columns:
                    df = format_headers(df)
                try:
                    df.to_excel(writer, sheet_name=clean_sheet_name(sheet_name), index=False)
                except Exception as e:
                    sheets_failed.append(sheet_name)

    if not sheets_failed:
        os.replace(temp_file, output_file)
    else:
        return sheets_failed

def run_summary_script(output_file, new_header):
    """Ejecuta resumen.py con el archivo de salida."""
    def ajustar_ancho_columnas(hoja):
        """Ajusta el ancho de las columnas según el contenido, incluyendo un margen."""
        for col in hoja.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Ajustar el ancho con un pequeño margen
            adjusted_width = (max_length + 4)  # Añadir margen de 4 caracteres
            hoja.column_dimensions[column].width = adjusted_width

    def cambiar_encabezado_y_crear_resumen(archivo, new_header):
        """Cambia el encabezado y crea una hoja de resumen."""
        if not archivo:
            raise ValueError("No se proporcionó ningún archivo.")

        # Carga el archivo Excel
        wb = load_workbook(archivo)

        # Cambiar encabezado en cada hoja
        old_header = 'InterviewID'
        print("Cambiando encabezado...")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                if cell.value == old_header:
                    cell.value = new_header
                    print(f"Encabezado cambiado en la hoja {sheet}")

        # Crear una nueva hoja para el resumen
        if 'Resumen' in wb.sheetnames:
            wb.remove(wb['Resumen'])
        hoja_resumen = wb.create_sheet("Resumen")

        # Encabezados para la hoja de resumen
        encabezados = ['HOJA', 'PREGUNTA', 'TIPO', 'ATRIBUTO', 'COD_OTROS', 'NIVEL_RPTA']
        hoja_resumen.append(encabezados)

        # Estilo para el encabezado
        celeste_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        borde = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        
        for cell in hoja_resumen[1]:
            cell.fill = celeste_claro
            cell.border = borde

        print("Hoja de resumen creada con encabezados.")

        # Expresión regular para identificar el patrón ._Número
        patron_tipo = re.compile(r'\._\d+$')

        for sheet in wb.sheetnames:
            if sheet.lower() not in ['verbatim', 'frequency', 'resumen']:  # Asegúrate de omitir las hojas especificadas
                print(f"Procesando la hoja: {sheet}")

                # Determinar el valor de PREGUNTA
                if '.' in sheet:
                    pregunta_base = sheet.split('.')[0]  # Antes del primer punto
                else:
                    pregunta_base = sheet  # Si no hay punto
                print(f"PREGUNTA_BASE: {pregunta_base}")

                # Determinar el valor de TIPO y COD_OTROS
                if patron_tipo.search(sheet):
                    tipo = 'SA'
                    cod_otros = '{' + re.search(r'_\d+$', sheet).group() + '}'
                    print(f"TIPO: SA, COD_OTROS: {cod_otros}")
                else:
                    tipo = 'A'
                    cod_otros = ''
                    print(f"TIPO: A, COD_OTROS: {cod_otros}")

                # Agregar sufijo _CODED si el TIPO es A
                if tipo == 'A':
                    pregunta_base += '_CODED'
                
                # Determinar el valor de ATRIBUTO y NIVEL_RPTA
                partes = re.split(r'(\.|_)', sheet)
                atributo = ''
                nivel_rpta = ''
                
                # Obtener ATRIBUTO
                if partes.count('.') > 1:  # Solo si hay más de un punto
                    for i, parte in enumerate(partes):
                        if parte == '_':
                            if i + 1 < len(partes) and partes[i + 1].isdigit():
                                atributo = '{_' + partes[i + 1] + '}'
                                break
                print(f"ATRIBUTO: {atributo}")

                # Obtener NIVEL_RPTA (lo que sigue después del segundo punto)
                puntos = [i for i, p in enumerate(partes) if p == '.']
                if len(puntos) > 1:  # Solo si hay más de un punto
                    segundo_punto = puntos[1]
                    if segundo_punto + 1 < len(partes):
                        nivel_rpta = partes[segundo_punto + 1]
                print(f"NIVEL_RPTA: {nivel_rpta}")

                hoja_resumen.append([sheet, pregunta_base, tipo, atributo, cod_otros, nivel_rpta])

        # Ajustar el ancho de las columnas
        ajustar_ancho_columnas(hoja_resumen)

        # Mover la hoja de resumen a la primera posición
        wb._sheets.sort(key=lambda sheet: (sheet.title != "Resumen", sheet.title))
        
        # Guarda los cambios
        wb.save(archivo)
        print(f"El archivo ha sido modificado y guardado como {archivo}")

        return archivo

    if os.path.isfile(output_file):
        cambiar_encabezado_y_crear_resumen(output_file, new_header)
