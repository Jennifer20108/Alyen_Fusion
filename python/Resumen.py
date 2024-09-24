import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import os

def ajustar_ancho_columnas(hoja):
    """Ajusta el ancho de las columnas según el contenido, incluyendo un margen."""
    for col in hoja.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Ajustar el ancho con un pequeño margen
        adjusted_width = (max_length + 4)  # Añadir margen de 4 caracteres
        hoja.column_dimensions[column].width = adjusted_width

def cambiar_encabezado_y_crear_resumen(archivo, new_header):
    """Cambia el encabezado y crea una hoja de resumen."""
    if not archivo:
        raise ValueError("No se proporcionó ningún archivo.")

    # Carga el archivo Excel
    wb = load_workbook(archivo)

    # Cambiar encabezado en cada hoja
    old_header = 'InterviewID'
    print("Cambiando encabezado...")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            if cell.value == old_header:
                cell.value = new_header
                print(f"Encabezado cambiado en la hoja {sheet}")

    # Crear una nueva hoja para el resumen
    if 'Resumen' in wb.sheetnames:
        wb.remove(wb['Resumen'])
    hoja_resumen = wb.create_sheet("Resumen")

    # Encabezados para la hoja de resumen
    encabezados = ['HOJA', 'PREGUNTA', 'TIPO', 'ATRIBUTO', 'COD_OTROS', 'NIVEL_RPTA']
    hoja_resumen.append(encabezados)

    # Estilo para el encabezado
    celeste_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    borde = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    
    for cell in hoja_resumen[1]:
        cell.fill = celeste_claro
        cell.border = borde

    print("Hoja de resumen creada con encabezados.")

    # Expresión regular para identificar el patrón ._Número
    patron_tipo = re.compile(r'\._\d+$')

    for sheet in wb.sheetnames:
        if sheet.lower() not in ['verbatim', 'frequency', 'resumen']:  # Asegúrate de omitir las hojas especificadas
            print(f"Procesando la hoja: {sheet}")

            # Determinar el valor de PREGUNTA
            if '.' in sheet:
                pregunta_base = sheet.split('.')[0]  # Antes del primer punto
            else:
                pregunta_base = sheet  # Si no hay punto
            print(f"PREGUNTA_BASE: {pregunta_base}")

            # Determinar el valor de TIPO y COD_OTROS
            if patron_tipo.search(sheet):
                tipo = 'SA'
                cod_otros = '{' + re.search(r'_\d+$', sheet).group() + '}'
                print(f"TIPO: SA, COD_OTROS: {cod_otros}")
            else:
                tipo = 'A'
                cod_otros = ''
                print(f"TIPO: A, COD_OTROS: {cod_otros}")

            # Agregar sufijo _CODED si el TIPO es A
            if tipo == 'A':
                pregunta_base += '_CODED'
            
            # Determinar el valor de ATRIBUTO y NIVEL_RPTA
            partes = re.split(r'(\.|_)', sheet)
            atributo = ''
            nivel_rpta = ''
            
            # Obtener ATRIBUTO
            if partes.count('.') > 1:  # Solo si hay más de un punto
                for i, parte in enumerate(partes):
                    if parte == '_':
                        if i + 1 < len(partes) and partes[i + 1].isdigit():
                            atributo = '{_' + partes[i + 1] + '}'
                            break
            print(f"ATRIBUTO: {atributo}")

            # Obtener NIVEL_RPTA (lo que sigue después del segundo punto)
            puntos = [i for i, p in enumerate(partes) if p == '.']
            if len(puntos) > 1:  # Solo si hay más de un punto
                segundo_punto = puntos[1]
                if segundo_punto + 1 < len(partes):
                    nivel_rpta = partes[segundo_punto + 1]
            print(f"NIVEL_RPTA: {nivel_rpta}")

            hoja_resumen.append([sheet, pregunta_base, tipo, atributo, cod_otros, nivel_rpta])

    # Ajustar el ancho de las columnas
    ajustar_ancho_columnas(hoja_resumen)

    # Mover la hoja de resumen a la primera posición
    wb._sheets.sort(key=lambda sheet: (sheet.title != "Resumen", sheet.title))
    
    # Guarda los cambios
    wb.save(archivo)
    print(f"El archivo ha sido modificado y guardado como {archivo}")

    return archivo
